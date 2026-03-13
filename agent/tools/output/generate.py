"""
Генерация Excel-отчёта.

При вызове generate_excel_report самостоятельно считает все 13 финансовых показателей
по всем годам, доступным в inputs_cache, и сохраняет результат в .xlsx.
Также включает лист "Анализ рисков" с цветовой маркировкой.
"""

from typing import Dict, List, Optional

from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from agent.indicators.compute import ALL_INDICATORS, IndicatorDef
from agent.indicators.risk_detector import LEVEL_CRITICAL, LEVEL_OK, LEVEL_WARNING, detect_all_risks
from agent.tools.output._shared import compute_all_indicators, get_years_or_error, make_output_path
from agent.tools.utils import EXCEL_OUTPUT_PATH_BASE

_SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FILL  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_FILL_RED     = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_YELLOW  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_GREEN   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_BOLD = Font(bold=True)
_DASH = "\u2014"  # —

_FMT_PERCENT = "0.00%"  # Excel умножает значение на 100 и добавляет %
_FMT_RATIO   = "0.00"   # коэффициенты — два знака после запятой

# Быстрый словарь: имя показателя -> is_percent
_IS_PERCENT: Dict[str, bool] = {ind.name: ind.is_percent for ind in ALL_INDICATORS}

_RISK_FILL = {
    LEVEL_CRITICAL: _FILL_RED,
    LEVEL_WARNING:  _FILL_YELLOW,
    LEVEL_OK:       _FILL_GREEN,
}


def _apply_value_formats(
    ws: Worksheet, row_num: int, col_start: int, n_cols: int, is_percent: bool
) -> None:
    """Применяет числовой формат ко всем ячейкам со значениями в строке."""
    fmt = _FMT_PERCENT if is_percent else _FMT_RATIO
    for col in range(col_start, col_start + n_cols):
        cell = ws.cell(row=row_num, column=col)
        if cell.value != _DASH and cell.value is not None:
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")


def _write_indicators_sheet(
    ws: Worksheet,
    years: List[str],
    results: Dict[str, Dict[str, Optional[float]]],
) -> None:
    """Заполняет лист: строки — показатели, сгруппированные по разделам; столбцы — годы."""

    year_col_start = 2

    ws.column_dimensions["A"].width = 45
    for i, year in enumerate(years):
        col_letter = chr(ord("A") + year_col_start - 1 + i)
        ws.column_dimensions[col_letter].width = 16

    header = ["Показатель"] + years
    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    seen_sections: set = set()
    used_names: set = set()

    def _append_indicator(ind: IndicatorDef, section: str) -> None:
        if section not in seen_sections:
            seen_sections.add(section)
            section_row = [section] + [""] * len(years)
            ws.append(section_row)
            for col_idx in range(1, len(years) + 2):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.font = _BOLD
                cell.fill = _SECTION_FILL

        year_values = results.get(ind.name, {})
        data_row = [ind.name] + [
            _DASH if year_values.get(y) is None else year_values.get(y)
            for y in years
        ]
        ws.append(data_row)
        _apply_value_formats(ws, ws.max_row, year_col_start, len(years), ind.is_percent)
        used_names.add(ind.name)

    for ind in ALL_INDICATORS:
        _append_indicator(ind, ind.section)

    leftover = [(name, vals) for name, vals in results.items() if name not in used_names]
    if leftover:
        section = "Прочее"
        seen_sections.discard(section)
        for name, vals in leftover:
            is_pct = _IS_PERCENT.get(name, False)
            ind_stub = IndicatorDef(name, section, lambda _y: None, is_pct)
            _append_indicator(ind_stub, section)
            row_num = ws.max_row
            for i, year in enumerate(years):
                v = vals.get(year)
                ws.cell(row=row_num, column=year_col_start + i).value = _DASH if v is None else v


def _write_risks_sheet(ws: Worksheet, years: List[str]) -> None:
    """
    Заполняет лист 'Анализ рисков': для каждого года — блок с флагами рисков,
    цветовая маркировка строк (красный/жёлтый/зелёный).
    """
    ws.column_dimensions["A"].width = 10   # уровень
    ws.column_dimensions["B"].width = 35   # показатель
    ws.column_dimensions["C"].width = 80   # сообщение

    header = ["Уровень", "Показатель", "Описание"]
    ws.append(header)
    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for year in years:
        # Строка-разделитель с годом
        ws.append([f"Год: {year}", "", ""])
        year_row = ws.max_row
        for col_idx in range(1, 4):
            cell = ws.cell(row=year_row, column=col_idx)
            cell.font = _BOLD
            cell.fill = _SECTION_FILL

        flags = detect_all_risks(year)
        for flag in flags:
            ws.append([flag.level, flag.indicator, flag.message])
            fill = _RISK_FILL.get(flag.level, _FILL_GREEN)
            for col_idx in range(1, 4):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.append(["", "", ""])  # пустая строка между годами


@tool(
    "generate_excel_report",
    description=(
        "Создаёт Excel (.xlsx) со всеми финансовыми показателями и анализом рисков. "
        "Считает все показатели самостоятельно по всем доступным годам. "
        "Параметры не нужны."
    ),
)
def generate_excel_report() -> str:
    """
    Рассчитывает все 13 финансовых показателей и анализ рисков по всем годам из inputs_cache
    и сохраняет в Excel (два листа).

    :return: путь к файлу или сообщение об ошибке
    """
    years = get_years_or_error()
    if isinstance(years, str):
        return years

    results = compute_all_indicators(years)
    file_path = make_output_path(EXCEL_OUTPUT_PATH_BASE, "financial_report", "xlsx")

    wb = Workbook()

    ws_indicators = wb.active
    ws_indicators.title = "Финансовые показатели"
    _write_indicators_sheet(ws_indicators, years, results)

    ws_risks = wb.create_sheet(title="Анализ рисков")
    _write_risks_sheet(ws_risks, years)

    wb.save(file_path)

    return f"Файл сохранён: {file_path}"
