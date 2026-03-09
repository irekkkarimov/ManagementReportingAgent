"""
Генерация Excel-отчёта с финансовыми показателями из кэша расчётов.
"""

import os
from datetime import datetime
from typing import Dict, List, Tuple

from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from agent.tools.finance.calculation_cache import get_indicators
from agent.tools.utils import EXCEL_OUTPUT_PATH_BASE

_SECTIONS: List[Tuple[str, List[str]]] = [
    ("Рентабельность", ["ROS", "ROA", "ROE", "Gross Margin", "Operating Margin"]),
    ("Оборачиваемость", ["Total Asset Turnover", "Inventory Turnover", "Receivables Turnover", "Payables Turnover"]),
    ("Финансовая устойчивость", ["Financial Stability Ratio"]),
    ("Ликвидность", ["Current Liquidity Ratio", "Quick Liquidity Ratio", "Cash Liquidity Ratio"]),
]


def _section_data(indicators: Dict[str, float]) -> Dict[str, Dict[str, float]]:
    return {
        title: {k: indicators[k] for k in keys if k in indicators}
        for title, keys in _SECTIONS
        if any(k in indicators for k in keys)
    }


def _write_sheet(ws: Worksheet, title: str, metrics: Dict[str, float]) -> None:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Показатель", "Значение"])
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    for name, value in metrics.items():
        ws.append([name, value])

    for col in ws.columns:
        width = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = width + 2


def _make_file_path() -> str:
    os.makedirs(EXCEL_OUTPUT_PATH_BASE, exist_ok=True)
    filename = f"financial_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return os.path.join(EXCEL_OUTPUT_PATH_BASE, filename)


@tool(
    "generate_excel_report",
    description=(
        "Создаёт Excel (xlsx) с финансовыми показателями из кэша расчётов. "
        "Данные берутся из последних вызовов инструментов calculate_*. "
        "Параметры не нужны. Если кэш пуст — сначала посчитай показатели."
    ),
)
def generate_excel_report() -> str:
    """
    :return: путь к сохранённому файлу или сообщение об ошибке
    """
    indicators = get_indicators()
    if not indicators:
        return (
            "Нет сохранённых расчётов. "
            "Сначала посчитай показатели с помощью calculate_* (ROS, ROA, ROE и т.д.)."
        )

    data = _section_data(indicators)
    file_path = _make_file_path()

    wb = Workbook()
    wb.remove(wb.active)
    for title, metrics in data.items():
        _write_sheet(wb.create_sheet(title=title[:31]), title, metrics)

    wb.save(file_path)
    return f"Файл сохранён: {file_path}"
